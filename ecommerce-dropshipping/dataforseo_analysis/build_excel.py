#!/usr/bin/env python3
"""Build final v3 Excel + report from results.json"""
import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
data = json.loads((HERE / "results.json").read_text())
products = data["products"]
variants = data["variants"]

# --- Build DataFrames ---
INITIAL_COLS = ['PRODUIT', 'NICHE', 'P. ACHAT', 'P. VENTE', 'MARGE BRUTE EST.',
                'SAISONNALITÉ', 'MOT-CLÉ', 'GOOGLE TRENDS', 'GOOGLE SHOPPING',
                'GOOGLE SEARCH', 'OFFRE + ANGLE MARKETING', 'RISQUE', 'NOTES',
                'SOURCE / ORIGINE', 'CHECK']
NEW_COLS = ['VOLUME_FR', 'CPC_EUR', 'COMPETITION_INDEX', 'NB_ADS_SERP',
            'GEANTS_TOP3', 'MOIS_PIC', 'RATIO_SAISONNALITE', 'TYPE_SAISON',
            'VOLUME_CUMULE_VARIANTES', 'NB_VARIANTES_VIABLES', 'SCORE_TOTAL',
            'TOP_ORGANIC', 'TOP_ADS']

all_cols = INITIAL_COLS + NEW_COLS

df_full = pd.DataFrame(products)
for c in all_cols:
    if c not in df_full.columns:
        df_full[c] = None
df_full = df_full[all_cols]

# Shortlist (sorted by score desc, score > 0)
df_shortlist = df_full[df_full["SCORE_TOTAL"] > 0].sort_values("SCORE_TOTAL", ascending=False).reset_index(drop=True)

# Variants detail
df_variants = pd.DataFrame(variants) if variants else pd.DataFrame(columns=["produit","keyword_principal","variante","volume","cpc_eur","competition_index"])
if not df_variants.empty:
    df_variants = df_variants.sort_values(["produit", "volume"], ascending=[True, False])

# --- Write Excel ---
output_path = HERE / "Tableau_Recherche_Produit_v3_dataforseo.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df_full.to_excel(writer, sheet_name="Tableau produits", index=False)
    df_shortlist.to_excel(writer, sheet_name="Shortlist", index=False)
    df_variants.to_excel(writer, sheet_name="Variantes_détail", index=False)

# --- Formatting ---
wb = load_workbook(output_path)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def color_for_score(s):
    if s is None: return None
    if s >= 60: return "C6EFCE"   # green
    if s >= 40: return "FFEB9C"   # yellow
    if s >= 30: return "FFD7A1"   # orange
    return "FFC7CE"               # red

for sname in ["Tableau produits", "Shortlist", "Variantes_détail"]:
    ws = wb[sname]
    # header
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36
    # body font
    max_col = ws.max_column
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    # column widths
    widths = {"PRODUIT": 38, "NICHE": 22, "P. ACHAT": 10, "P. VENTE": 10,
              "MARGE BRUTE EST.": 16, "SAISONNALITÉ": 24, "MOT-CLÉ": 30,
              "GOOGLE TRENDS": 12, "GOOGLE SHOPPING": 12, "GOOGLE SEARCH": 12,
              "OFFRE + ANGLE MARKETING": 50, "RISQUE": 18, "NOTES": 40,
              "SOURCE / ORIGINE": 22, "CHECK": 8, "VOLUME_FR": 12,
              "CPC_EUR": 10, "COMPETITION_INDEX": 14, "NB_ADS_SERP": 12,
              "GEANTS_TOP3": 13, "MOIS_PIC": 12, "RATIO_SAISONNALITE": 14,
              "TYPE_SAISON": 18, "VOLUME_CUMULE_VARIANTES": 16,
              "NB_VARIANTES_VIABLES": 14, "SCORE_TOTAL": 12,
              "TOP_ORGANIC": 50, "TOP_ADS": 50, "produit": 38,
              "keyword_principal": 28, "variante": 30, "volume": 10,
              "cpc_eur": 10, "competition_index": 14}
    for i in range(1, max_col + 1):
        h = ws.cell(row=1, column=i).value
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 14)
    # freeze
    ws.freeze_panes = "A2"

    # color score column
    if "SCORE_TOTAL" in [ws.cell(row=1, column=i).value for i in range(1, max_col + 1)]:
        score_col = [ws.cell(row=1, column=i).value for i in range(1, max_col + 1)].index("SCORE_TOTAL") + 1
        for r in range(2, max_row + 1):
            cell = ws.cell(row=r, column=score_col)
            color = color_for_score(cell.value)
            if color:
                cell.fill = PatternFill("solid", start_color=color)
                cell.font = Font(name="Arial", size=11, bold=True)

# --- Build Rapport sheet ---
ws_r = wb.create_sheet("Rapport", 0)
ws_r.column_dimensions["A"].width = 6
ws_r.column_dimensions["B"].width = 50
ws_r.column_dimensions["C"].width = 12
ws_r.column_dimensions["D"].width = 12
ws_r.column_dimensions["E"].width = 16
ws_r.column_dimensions["F"].width = 14
ws_r.column_dimensions["G"].width = 14
ws_r.column_dimensions["H"].width = 22

def title(row, text, size=14):
    c = ws_r.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", size=size, bold=True, color="1F4E78")
    ws_r.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    return row + 1

def header_row(row, headers):
    for i, h in enumerate(headers):
        c = ws_r.cell(row=row, column=i+1, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws_r.row_dimensions[row].height = 32
    return row + 1

def write_row(row, values, score_col_idx=None):
    for i, v in enumerate(values):
        c = ws_r.cell(row=row, column=i+1, value=v)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BORDER
        if score_col_idx is not None and i == score_col_idx and isinstance(v, (int, float)):
            color = color_for_score(v)
            if color:
                c.fill = PatternFill("solid", start_color=color)
                c.font = Font(name="Arial", size=10, bold=True)
    return row + 1

row = 1
row = title(row, "RAPPORT DATAFORSEO — Recherche produits e-commerce", 16)
ws_r.cell(row=row, column=1, value=f"Pipeline : Google Ads Search Volume + Keyword Suggestions + SERP (FR, fr)").font = Font(name="Arial", size=10, italic=True)
ws_r.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 1
ws_r.cell(row=row, column=1, value=f"189 appels API · 46 produits analysés · {len(variants)} variantes viables retenues").font = Font(name="Arial", size=10, italic=True)
ws_r.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 2

# TOP 10
row = title(row, "🏆 TOP 10 — Produits par SCORE_TOTAL", 13)
row = header_row(row, ["#", "PRODUIT", "SCORE", "VOL FR", "CPC €", "COMP. IDX", "GÉANTS T3", "RISQUE"])
top10 = df_shortlist.head(10)
for i, (_, p) in enumerate(top10.iterrows(), 1):
    row = write_row(row, [i, p["PRODUIT"], int(p["SCORE_TOTAL"]),
                          p["VOLUME_FR"], p["CPC_EUR"], p["COMPETITION_INDEX"],
                          p["GEANTS_TOP3"], p["RISQUE"]],
                    score_col_idx=2)
row += 1

# TOP 5 Faible risque + bon score
risk_filter = df_shortlist[
    df_shortlist["RISQUE"].astype(str).str.lower().str.contains("faible", na=False) &
    (df_shortlist["SCORE_TOTAL"] >= 30)
].head(5)
row = title(row, "🛡️ TOP 5 — Faible risque + bon score (prioritaire pour test)", 13)
row = header_row(row, ["#", "PRODUIT", "SCORE", "VOL FR", "CPC €", "COMP. IDX", "GÉANTS T3", "NOTES"])
for i, (_, p) in enumerate(risk_filter.iterrows(), 1):
    notes = (str(p.get("NOTES", "") or ""))[:90]
    row = write_row(row, [i, p["PRODUIT"], int(p["SCORE_TOTAL"]),
                          p["VOLUME_FR"], p["CPC_EUR"], p["COMPETITION_INDEX"],
                          p["GEANTS_TOP3"], notes], score_col_idx=2)
row += 1

# À éliminer
elim = df_full[df_full["SCORE_TOTAL"] < 30].sort_values("SCORE_TOTAL")
row = title(row, f"❌ À ÉLIMINER — Score < 30 ({len(elim)} produits)", 13)
row = header_row(row, ["#", "PRODUIT", "SCORE", "VOL FR", "CPC €", "GÉANTS T3", "RISQUE", "RAISON"])
for i, (_, p) in enumerate(elim.iterrows(), 1):
    vol = p.get("VOLUME_FR")
    reasons = []
    if vol is None or vol == 0: reasons.append("aucun volume")
    elif vol < 500: reasons.append("vol. très faible")
    if p.get("GEANTS_TOP3") == "Oui": reasons.append("géants top 3")
    cpc = p.get("CPC_EUR")
    if cpc and cpc > 1.50: reasons.append(f"CPC élevé {cpc}€")
    comp = p.get("COMPETITION_INDEX")
    if comp and comp > 70: reasons.append("compétition forte")
    reason = " · ".join(reasons) if reasons else "score insuffisant"
    row = write_row(row, [i, p["PRODUIT"], int(p["SCORE_TOTAL"]) if pd.notna(p["SCORE_TOTAL"]) else 0,
                          vol, cpc, p.get("GEANTS_TOP3"), p.get("RISQUE"), reason], score_col_idx=2)

ws_r.freeze_panes = "A2"

# Save
wb.save(output_path)
print(f"[✓] Saved {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Top 10: {len(top10)} | Faible risque top5: {len(risk_filter)} | À éliminer: {len(elim)}")
