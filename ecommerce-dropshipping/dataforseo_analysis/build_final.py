#!/usr/bin/env python3
"""
Build single-sheet Excel preserving original design:
- Blue header row (1F4E78)
- Black month-bar separator rows
- DataForSEO columns appended
- Recommendations injected into NOTES column
"""
import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
INPUT_FILE = "/Users/Hakim/Downloads/Tableau_Recherche_Produit_v2.xlsx"
OUTPUT_FILE = HERE / "Tableau_Recherche_Produit_v3_dataforseo.xlsx"

# Load source structure (with month rows)
src_df = pd.read_excel(INPUT_FILE, sheet_name="Tableau produits")

# Load DataForSEO results indexed by PRODUIT
results = json.loads((HERE / "results.json").read_text())
by_product = {p["PRODUIT"]: p for p in results["products"]}

# Columns (original 15 + DataForSEO additions)
ORIG_COLS = list(src_df.columns)  # 15 columns
DFSEO_COLS = ["VOLUME_FR", "CPC_EUR", "COMPETITION_INDEX", "NB_ADS_SERP",
              "GEANTS_TOP3", "MOIS_PIC", "RATIO_SAISONNALITE", "TYPE_SAISON",
              "VOLUME_CUMULE_VARIANTES", "NB_VARIANTES_VIABLES", "SCORE_TOTAL",
              "TOP_ORGANIC", "TOP_ADS"]
ALL_COLS = ORIG_COLS + DFSEO_COLS

MONTHS = {'JANVIER','FÉVRIER','MARS','AVRIL','MAI','JUIN','JUILLET','AOÛT',
          'SEPTEMBRE','OCTOBRE','NOVEMBRE','DÉCEMBRE'}


def build_recommendation(d):
    """Build commentary for NOTES based on score + signals."""
    if not d:
        return None
    s = d.get("SCORE_TOTAL", 0) or 0
    vol = d.get("VOLUME_FR")
    cpc = d.get("CPC_EUR")
    comp = d.get("COMPETITION_INDEX")
    giants = d.get("GEANTS_TOP3") == "Oui"
    season = d.get("TYPE_SAISON")
    ratio = d.get("RATIO_SAISONNALITE")
    n_var = d.get("NB_VARIANTES_VIABLES") or 0

    parts = []
    # Verdict
    if s >= 60:
        parts.append("[DataForSEO] ✅ À TESTER — bon profil global.")
    elif s >= 40:
        parts.append("[DataForSEO] 🟡 Potentiel modéré, à challenger.")
    elif s >= 30:
        parts.append("[DataForSEO] 🟠 Limite — niche d'audience requise.")
    else:
        parts.append("[DataForSEO] ❌ Score faible, à écarter sauf angle différenciant.")

    # Signals
    if vol is None or vol == 0:
        parts.append("Aucun volume FR détecté sur le keyword.")
    elif vol < 500:
        parts.append(f"Volume FR faible ({vol}/mois).")
    elif vol >= 5000:
        parts.append(f"Bon volume ({vol}/mois).")

    if cpc and cpc > 1.50:
        parts.append(f"CPC élevé ({cpc}€) — Ads coûteux.")
    elif cpc and 0.30 <= cpc <= 1.00:
        parts.append(f"CPC sain ({cpc}€).")

    if giants:
        parts.append("⚠️ Géants en TOP 3 SERP — concurrence frontale.")

    if season == "saisonnier fort" and ratio:
        parts.append(f"Saisonnalité forte (ratio {ratio}).")

    if n_var >= 5:
        parts.append(f"{n_var} variantes longue-traîne viables.")

    return " ".join(parts)


# Build workbook
wb = Workbook()
ws = wb.active
ws.title = "Tableau produits"

# Header
for i, col in enumerate(ALL_COLS, 1):
    ws.cell(row=1, column=i, value=col)

# Data rows preserving order (month bars + product rows)
out_row = 2
month_row_idx = []
for _, src_row in src_df.iterrows():
    produit = src_row["PRODUIT"]
    if produit in MONTHS:
        ws.cell(row=out_row, column=1, value=produit)
        month_row_idx.append(out_row)
        out_row += 1
        continue
    if pd.isna(produit):
        out_row += 1
        continue

    # original columns
    for ci, col in enumerate(ORIG_COLS, 1):
        v = src_row[col]
        if pd.isna(v):
            v = None
        ws.cell(row=out_row, column=ci, value=v)

    # DFSEO data
    d = by_product.get(produit)
    if d:
        for ci, col in enumerate(DFSEO_COLS, len(ORIG_COLS) + 1):
            ws.cell(row=out_row, column=ci, value=d.get(col))

        # Inject DataForSEO recommendation into NOTES (col 13)
        existing_notes = src_row["NOTES"] if pd.notna(src_row["NOTES"]) else ""
        reco = build_recommendation(d)
        merged_notes = f"{existing_notes}\n\n— {reco}" if existing_notes else reco
        ws.cell(row=out_row, column=ORIG_COLS.index("NOTES") + 1, value=merged_notes)

    out_row += 1

# --- Styling ---
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
MONTH_FILL = PatternFill("solid", start_color="000000")
MONTH_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

n_cols = len(ALL_COLS)
n_rows = out_row - 1

# Header style
for c in range(1, n_cols + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
ws.row_dimensions[1].height = 36

# Month rows: span all columns, black bg, white bold centered
for mr in month_row_idx:
    ws.merge_cells(start_row=mr, start_column=1, end_row=mr, end_column=n_cols)
    cell = ws.cell(row=mr, column=1)
    cell.fill = MONTH_FILL
    cell.font = MONTH_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[mr].height = 26

# Body styling + score coloring
def color_for_score(s):
    if s is None: return None
    try: s = int(s)
    except: return None
    if s >= 60: return "C6EFCE"
    if s >= 40: return "FFEB9C"
    if s >= 30: return "FFD7A1"
    if s > 0:   return "FFC7CE"
    return None

score_col_idx = ALL_COLS.index("SCORE_TOTAL") + 1
geants_col_idx = ALL_COLS.index("GEANTS_TOP3") + 1
notes_col_idx = ALL_COLS.index("NOTES") + 1

for r in range(2, n_rows + 1):
    if r in month_row_idx:
        continue
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=r, column=c)
        if cell.value is None and c > 1 and ws.cell(row=r, column=1).value is None:
            continue
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER
    # Color score cell
    sc = ws.cell(row=r, column=score_col_idx)
    if sc.value is not None:
        col = color_for_score(sc.value)
        if col:
            sc.fill = PatternFill("solid", start_color=col)
            sc.font = Font(name="Calibri", size=11, bold=True)
    # Color GEANTS_TOP3 cell
    gc = ws.cell(row=r, column=geants_col_idx)
    if gc.value == "Oui":
        gc.fill = PatternFill("solid", start_color="FFC7CE")
        gc.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
    elif gc.value == "Non":
        gc.fill = PatternFill("solid", start_color="C6EFCE")
        gc.font = Font(name="Calibri", size=10, bold=True, color="006100")

# Column widths
widths = {
    "PRODUIT": 38, "NICHE": 24, "P. ACHAT": 12, "P. VENTE": 12,
    "MARGE BRUTE EST.": 16, "SAISONNALITÉ": 22, "MOT-CLÉ": 30,
    "GOOGLE TRENDS": 14, "GOOGLE SHOPPING": 14, "GOOGLE SEARCH": 14,
    "OFFRE + ANGLE MARKETING": 50, "RISQUE": 22, "NOTES": 60,
    "SOURCE / ORIGINE": 26, "CHECK": 10,
    "VOLUME_FR": 11, "CPC_EUR": 10, "COMPETITION_INDEX": 13,
    "NB_ADS_SERP": 11, "GEANTS_TOP3": 13, "MOIS_PIC": 12,
    "RATIO_SAISONNALITE": 14, "TYPE_SAISON": 18,
    "VOLUME_CUMULE_VARIANTES": 14, "NB_VARIANTES_VIABLES": 14,
    "SCORE_TOTAL": 13, "TOP_ORGANIC": 48, "TOP_ADS": 48,
}
for i, c in enumerate(ALL_COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)

ws.freeze_panes = "A2"

# Add auto-filter on header
ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"

wb.save(OUTPUT_FILE)
print(f"[✓] Saved {OUTPUT_FILE}")
print(f"   Sheet: {ws.title} — {n_rows} rows × {n_cols} cols")
print(f"   Month bars: {len(month_row_idx)}")
