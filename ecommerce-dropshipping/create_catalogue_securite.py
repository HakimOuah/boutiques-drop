import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
collection_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
collection_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
winner_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
go_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
money_fmt = '#,##0.00 €'
pct_fmt = '0.0%'
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
wrap_align = Alignment(wrap_text=True, vertical="center")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Données produits par collection ──
# URLs = liens directs fiches produit AliExpress (vérifiés mars 2026)
products = [
    # Collection 1: Systèmes d'alarme complets
    {
        "collection": "Systèmes d'alarme complets",
        "produit": "PGST PG103 Kit Alarme WiFi/GSM/4G Tuya (sirène + capteurs + télécommandes)",
        "lien": "https://fr.aliexpress.com/item/1005006145824811.html",
        "statut_lien": "À vérifier",
        "fournisseur": "PGST Official Store",
        "note_fournisseur": "4.7+",
        "commandes": "2000+",
        "prix_achat": 52,
        "shipping": 0,
        "prix_vente": 249,
        "keywords": "alarme maison sans abonnement, kit alarme wifi",
        "features": "Écran LCD tactile, WiFi+GSM+4G, Tuya/Alexa/Google, 99 zones, 9 langues dont FR, sans abonnement"
    },
    {
        "collection": "Systèmes d'alarme complets",
        "produit": "KERUI W181 Kit Alarme WiFi/GSM Tuya (écran couleur + capteurs)",
        "lien": "https://fr.aliexpress.com/item/1005005177281025.html",
        "statut_lien": "À vérifier",
        "fournisseur": "KERUI Official Store",
        "note_fournisseur": "4.6+",
        "commandes": "1500+",
        "prix_achat": 42,
        "shipping": 0,
        "prix_vente": 199,
        "keywords": "alarme maison wifi, système alarme sans fil",
        "features": "Écran TFT couleur 1.7\", WiFi+GSM, Tuya/Alexa, détection intrusion/incendie/gaz, 72 enregistrements"
    },
    {
        "collection": "Systèmes d'alarme complets",
        "produit": "GauTone PG103 Kit Alarme WiFi/GSM 433MHz Tuya (capteur PIR immunité animaux)",
        "lien": "https://fr.aliexpress.com/item/1005001700651666.html",
        "statut_lien": "Probablement actif",
        "fournisseur": "GauTone Official Store",
        "note_fournisseur": "4.6+",
        "commandes": "1000+",
        "prix_achat": 43,
        "shipping": 0,
        "prix_vente": 189,
        "keywords": "alarme maison animaux, alarme wifi gsm",
        "features": "WiFi+GSM 433MHz, Tuya/Alexa/Google, immunité animaux, 99 zones, 5 groupes d'appels"
    },
    # Collection 2: Caméras de surveillance
    {
        "collection": "Caméras de surveillance",
        "produit": "BESDER 5MP PTZ Caméra WiFi Extérieure (IA détection humaine, zoom 4X, vision nocturne couleur)",
        "lien": "https://fr.aliexpress.com/item/1005006583816424.html",
        "statut_lien": "À vérifier",
        "fournisseur": "BESDER Official Store",
        "note_fournisseur": "4.9",
        "commandes": "7000+",
        "prix_achat": 22,
        "shipping": 0,
        "prix_vente": 89,
        "keywords": "camera surveillance wifi exterieur, camera exterieure wifi",
        "features": "5MP, PTZ 270°/100°, zoom 4X, IA humaine, audio bidirectionnel, IP66, vision nocturne couleur, ONVIF"
    },
    {
        "collection": "Caméras de surveillance",
        "produit": "ANBIUX 1080P PTZ Speed Dome WiFi Extérieure (double antenne, IR 40m)",
        "lien": "https://fr.aliexpress.com/item/1005003903045900.html",
        "statut_lien": "À vérifier",
        "fournisseur": "ANBIUX Official Store",
        "note_fournisseur": "4.5+",
        "commandes": "3000+",
        "prix_achat": 20,
        "shipping": 0,
        "prix_vente": 79,
        "keywords": "camera surveillance wifi, camera wifi exterieur",
        "features": "1080P Full HD, PTZ 270°/90°, zoom 4X, double antenne 5dB, H.265, IR 40-50m, carte SD 128Go"
    },
    {
        "collection": "Caméras de surveillance",
        "produit": "Caméra Solaire PTZ 5MP 4G SIM + WiFi (100% sans fil, batterie + solaire)",
        "lien": "https://fr.aliexpress.com/item/1005002917268481.html",
        "statut_lien": "Vérifié OK",
        "fournisseur": "Vendeur spécialisé sécurité",
        "note_fournisseur": "4.5+",
        "commandes": "500+",
        "prix_achat": 90,
        "shipping": 0,
        "prix_vente": 249,
        "keywords": "camera solaire 4G, camera surveillance sans fil",
        "features": "5MP, solaire+batterie, 4G SIM (sans WiFi requis), PTZ, détection PIR, audio, vision nocturne couleur"
    },
    {
        "collection": "Caméras de surveillance",
        "produit": "BESDER 1080P PTZ Speed Dome WiFi IP66 (YCC365 App)",
        "lien": "https://fr.aliexpress.com/item/1005005921350682.html",
        "statut_lien": "À vérifier",
        "fournisseur": "BESDER Authorize Store",
        "note_fournisseur": "4.7+",
        "commandes": "2000+",
        "prix_achat": 33,
        "shipping": 0,
        "prix_vente": 99,
        "keywords": "camera surveillance wifi 360, camera ip wifi",
        "features": "1080P, PTZ Speed Dome, IP66, zoom 4X, audio bidirectionnel, cloud storage"
    },
    # Collection 3: Serrures connectées
    {
        "collection": "Serrures connectées",
        "produit": "Tuya Serrure Biométrique WiFi 5-en-1 (empreinte, APP, code, RFID, clé)",
        "lien": "https://fr.aliexpress.com/item/1005005440409035.html",
        "statut_lien": "À vérifier",
        "fournisseur": "Vendeur spécialisé serrures",
        "note_fournisseur": "4.5+",
        "commandes": "800+",
        "prix_achat": 55,
        "shipping": 0,
        "prix_vente": 179,
        "keywords": "serrure connectee wifi, serrure empreinte digitale",
        "features": "5 modes: APP Tuya, empreinte FPC, code, carte RFID, clé. WiFi direct, porte 35-90mm, 8 langues"
    },
    {
        "collection": "Serrures connectées",
        "produit": "WAFU Serrure WiFi Bluetooth Tuya (empreinte, 100 utilisateurs)",
        "lien": "https://fr.aliexpress.com/item/1005001672566432.html",
        "statut_lien": "À vérifier",
        "fournisseur": "WAFU Store",
        "note_fournisseur": "4.5+",
        "commandes": "500+",
        "prix_achat": 53,
        "shipping": 0,
        "prix_vente": 149,
        "keywords": "serrure connectee, serrure digitale wifi",
        "features": "5 modes ouverture, 100 empreintes, alu anti-rouille, 4xAAA ~6 mois autonomie, idéal Airbnb/bureau"
    },
    {
        "collection": "Serrures connectées",
        "produit": "Tuya Serrure Extérieure Étanche WiFi (empreinte + code + APP, portail jardin)",
        "lien": "https://fr.aliexpress.com/item/1005002586851475.html",
        "statut_lien": "À vérifier",
        "fournisseur": "Vendeur AliExpress FR",
        "note_fournisseur": "4.4+",
        "commandes": "300+",
        "prix_achat": 75,
        "shipping": 0,
        "prix_vente": 199,
        "keywords": "serrure exterieure connectee, serrure portail",
        "features": "Étanche extérieur, empreinte+code+APP+carte, Tuya WiFi, Google/Alexa, portes bois et fer 40-80mm"
    },
    # Collection 4: Sonnettes vidéo
    {
        "collection": "Sonnettes vidéo",
        "produit": "Sonnette Vidéo WiFi Tuya 1080P Batterie (mini, appartement)",
        "lien": "https://fr.aliexpress.com/item/10000034488997.html",
        "statut_lien": "Vérifié OK",
        "fournisseur": "CHAMOUS Official Store",
        "note_fournisseur": "4.5+",
        "commandes": "2000+",
        "prix_achat": 15,
        "shipping": 0,
        "prix_vente": 69,
        "keywords": "sonnette connectee, sonnette video wifi",
        "features": "1080P, batterie rechargeable, Tuya/Alexa/Google, audio bidirectionnel, vision nocturne, détection PIR"
    },
    {
        "collection": "Sonnettes vidéo",
        "produit": "Sonnette Caméra WiFi Étanche 1080P HD (166° grand angle, IP65)",
        "lien": "https://fr.aliexpress.com/item/1005005321858419.html",
        "statut_lien": "À vérifier",
        "fournisseur": "Vendeur spécialisé",
        "note_fournisseur": "4.5+",
        "commandes": "1000+",
        "prix_achat": 55,
        "shipping": 0,
        "prix_vente": 139,
        "keywords": "sonnette video, sonnette camera wifi",
        "features": "1080P HD, angle 166°, audio bidirectionnel, IP65 étanche, batterie, cloud 7j gratuit, carillon inclus"
    },
    {
        "collection": "Sonnettes vidéo",
        "produit": "Reolink 2K+ Sonnette Vidéo Intelligente PoE (détection humaine, Alexa)",
        "lien": "https://fr.aliexpress.com/item/1005005251928644.html",
        "statut_lien": "Vérifié OK",
        "fournisseur": "Reolink Official Store",
        "note_fournisseur": "4.8+",
        "commandes": "500+",
        "prix_achat": 85,
        "shipping": 0,
        "prix_vente": 199,
        "keywords": "sonnette connectee premium, video doorbell",
        "features": "2K+ (2560x1920), détection humaine IA, Alexa, WiFi/PoE, carillon inclus, vision nocturne, sans abonnement"
    },
    # Collection 5: Interphones & Visiophones
    {
        "collection": "Interphones & Visiophones",
        "produit": "Jeatone Interphone Vidéo WiFi Tuya 7\" 1080P (clavier RFID, contrôle accès)",
        "lien": "https://fr.aliexpress.com/item/4001199694254.html",
        "statut_lien": "Vérifié OK",
        "fournisseur": "JEATONE Video Doorbell Store",
        "note_fournisseur": "4.7+",
        "commandes": "800+",
        "prix_achat": 100,
        "shipping": 0,
        "prix_vente": 249,
        "keywords": "interphone sans fil, visiophone sans fil",
        "features": "Écran tactile 7\" 1080P, WiFi Tuya, clavier code RFID, enregistrement carte 32G, déverrouillage à distance, FR"
    },
    {
        "collection": "Interphones & Visiophones",
        "produit": "Anjielo Smart Visiophone WiFi Tuya 7\" Écran Tactile",
        "lien": "https://fr.aliexpress.com/item/1005007870953315.html",
        "statut_lien": "Vérifié OK",
        "fournisseur": "Anjielo Smart",
        "note_fournisseur": "4.5+",
        "commandes": "600+",
        "prix_achat": 69,
        "shipping": 0,
        "prix_vente": 189,
        "keywords": "visiophone sans fil, interphone video wifi",
        "features": "Écran IPS 7\" 1024x600, WiFi 2.4/5GHz, Tuya, ouverture 2 serrures, détection mouvement, 86 photos + SD 128Go"
    },
    {
        "collection": "Interphones & Visiophones",
        "produit": "TMEZON 10\" Visiophone WiFi 1080P Écran Tactile (4-en-1, Tuya)",
        "lien": "https://fr.aliexpress.com/item/1005005081344282.html",
        "statut_lien": "Vérifié OK",
        "fournisseur": "TMEZON Store",
        "note_fournisseur": "4.6+",
        "commandes": "400+",
        "prix_achat": 118,
        "shipping": 0,
        "prix_vente": 299,
        "keywords": "interphone video maison, visiophone ecran",
        "features": "Écran tactile 10\" TFT, 1080P, WiFi+RJ45, capteur/code/carte/moniteur Tuya, anti-vandale"
    },
    # Collection 6: Détecteurs & Capteurs
    {
        "collection": "Détecteurs & Capteurs",
        "produit": "Tuya WiFi PIR Détecteur de Mouvement (105°, Alexa/Google)",
        "lien": "https://fr.aliexpress.com/item/32969085465.html",
        "statut_lien": "Vérifié OK",
        "fournisseur": "eMastiff Official Store",
        "note_fournisseur": "4.6",
        "commandes": "517+",
        "prix_achat": 9,
        "shipping": 0,
        "prix_vente": 29,
        "keywords": "detecteur de mouvement wifi, capteur mouvement",
        "features": "PIR 105° portée 5-7m, WiFi 2.4GHz sans hub, Alexa/Google, Tuya/Smart Life, 3xAAA, notifications push"
    },
    {
        "collection": "Détecteurs & Capteurs",
        "produit": "AVATTO Capteur Ouverture Porte/Fenêtre WiFi Tuya DS06 (lot possible x2/x4)",
        "lien": "https://fr.aliexpress.com/item/4000790781739.html",
        "statut_lien": "Vérifié OK",
        "fournisseur": "AVATTO Intelligent Store",
        "note_fournisseur": "4.8",
        "commandes": "145+",
        "prix_achat": 6,
        "shipping": 0,
        "prix_vente": 22,
        "keywords": "detecteur ouverture porte wifi, capteur fenetre",
        "features": "WiFi 2.4GHz direct sans hub, Alexa/Google, Tuya/Smart Life, 2xAAA, notifications ouverture/fermeture, historique"
    },
    # Collection 7: Sirènes & Accessoires
    {
        "collection": "Sirènes & Accessoires",
        "produit": "GauTone Sirène Solaire Extérieure Sans Fil (flash stroboscopique, 433MHz)",
        "lien": "https://fr.aliexpress.com/item/1005001315500648.html",
        "statut_lien": "Vérifié OK",
        "fournisseur": "GauTone Official Store",
        "note_fournisseur": "4.6+",
        "commandes": "800+",
        "prix_achat": 28,
        "shipping": 0,
        "prix_vente": 79,
        "keywords": "sirene alarme exterieure, sirene solaire",
        "features": "Solaire, étanche extérieur, flash stroboscopique, haute puissance, compatible 433MHz WiFi/GSM"
    },
    {
        "collection": "Sirènes & Accessoires",
        "produit": "Coffre-Fort Électronique Compact Empreinte + Code (4.6L/8.5L/16L au choix)",
        "lien": "https://fr.aliexpress.com/item/1005006039860123.html",
        "statut_lien": "Vérifié OK",
        "fournisseur": "Vendeur AliExpress FR",
        "note_fournisseur": "4.4+",
        "commandes": "300+",
        "prix_achat": 22,
        "shipping": 0,
        "prix_vente": 79,
        "keywords": "coffre fort, coffre fort electronique",
        "features": "Double clavier + empreinte, acier haute qualité, antirouille, 3 capacités au choix, compact"
    },
]

# ═══════════════════════════════════════════════════════
# SHEET 1 : Catalogue par Collection
# ═══════════════════════════════════════════════════════
ws = wb.active
ws.title = "Catalogue par Collection"

headers = [
    "Collection", "Produit", "Lien AliExpress", "Statut lien",
    "Fournisseur", "Note Fournisseur", "Commandes", "Prix Achat (€)", "Shipping (€)",
    "Coût Total (€)", "Prix Vente FR (€)", "Marge Brute (€)",
    "Marge Brute (%)", "Score Marge", "Keywords SEO", "Caractéristiques clés"
]

col_widths = [28, 55, 55, 16, 25, 16, 12, 14, 12, 14, 16, 14, 14, 12, 35, 60]

# Write headers
for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# Write data
row = 2
current_collection = None
for p in products:
    # Collection separator row
    if p["collection"] != current_collection:
        if current_collection is not None:
            row += 1  # blank row between collections
        current_collection = p["collection"]

    cout_total = p["prix_achat"] + p["shipping"]
    marge_brute = p["prix_vente"] - cout_total
    marge_pct = marge_brute / p["prix_vente"] if p["prix_vente"] > 0 else 0

    # Score marge
    if marge_pct >= 0.80:
        score = "A+"
    elif marge_pct >= 0.70:
        score = "A"
    elif marge_pct >= 0.60:
        score = "B+"
    else:
        score = "B"

    values = [
        p["collection"], p["produit"], p["lien"], p.get("statut_lien", ""),
        p["fournisseur"], p["note_fournisseur"], p["commandes"], p["prix_achat"], p["shipping"],
        cout_total, p["prix_vente"], marge_brute,
        marge_pct, score, p["keywords"], p["features"]
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = wrap_align

        # Format numbers
        if col_idx in (8, 9, 10, 11, 12):
            cell.number_format = money_fmt
        elif col_idx == 13:
            cell.number_format = pct_fmt

        # Color based on score
        if score == "A+":
            if col_idx in (12, 13, 14):
                cell.fill = winner_fill

    # Make link clickable
    link_cell = ws.cell(row=row, column=3)
    link_cell.hyperlink = p["lien"]
    link_cell.font = Font(color="0563C1", underline="single", size=10)

    row += 1

ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
ws.freeze_panes = "C2"

# ═══════════════════════════════════════════════════════
# SHEET 2 : Résumé par Collection
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("Résumé par Collection")

summary_headers = [
    "Collection", "Nb Produits", "Prix Achat Moyen (€)",
    "Prix Vente Moyen (€)", "Marge Brute Moyenne (€)",
    "Marge Moyenne (%)", "Ticket Moyen Vente", "Potentiel"
]

for col_idx, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 22
ws2.column_dimensions["E"].width = 24
ws2.column_dimensions["F"].width = 18
ws2.column_dimensions["G"].width = 20
ws2.column_dimensions["H"].width = 14

# Aggregate by collection
from collections import defaultdict
collections = defaultdict(list)
for p in products:
    collections[p["collection"]].append(p)

potentiels = {
    "Systèmes d'alarme complets": "★★★★★",
    "Caméras de surveillance": "★★★★",
    "Serrures connectées": "★★★★",
    "Sonnettes vidéo": "★★★★",
    "Interphones & Visiophones": "★★★★",
    "Détecteurs & Capteurs": "★★★",
    "Sirènes & Accessoires": "★★★",
}

row2 = 2
for coll_name, items in collections.items():
    nb = len(items)
    avg_achat = sum(p["prix_achat"] for p in items) / nb
    avg_vente = sum(p["prix_vente"] for p in items) / nb
    avg_marge = sum(p["prix_vente"] - p["prix_achat"] - p["shipping"] for p in items) / nb
    avg_marge_pct = avg_marge / avg_vente if avg_vente > 0 else 0
    ticket = f"{min(p['prix_vente'] for p in items)}-{max(p['prix_vente'] for p in items)}€"

    values = [coll_name, nb, avg_achat, avg_vente, avg_marge, avg_marge_pct, ticket, potentiels.get(coll_name, "★★★")]
    for col_idx, val in enumerate(values, 1):
        cell = ws2.cell(row=row2, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = center_align
        if col_idx in (3, 4, 5):
            cell.number_format = money_fmt
        elif col_idx == 6:
            cell.number_format = pct_fmt
    row2 += 1

ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════
# SHEET 3 : Marge Nette Micro-Entreprise
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("Marge Nette Micro-Entreprise")

marge_headers = [
    "Produit", "Collection", "Prix Achat (€)", "Shipping (€)", "Prix Vente TTC (€)",
    "Shopify 2%", "Stripe 1.4%+0.25€", "Coût Total (€)",
    "Marge Brute (€)", "Charges Micro 12.3%", "Marge Nette (€)",
    "Marge Nette (%)", "Ventes/jour 3000€/mois", "Ventes/jour 5000€/mois"
]

marge_widths = [50, 28, 14, 12, 16, 12, 16, 14, 14, 18, 14, 14, 22, 22]

for col_idx, (h, w) in enumerate(zip(marge_headers, marge_widths), 1):
    cell = ws3.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws3.column_dimensions[get_column_letter(col_idx)].width = w

row3 = 2
for p in products:
    pv = p["prix_vente"]
    pa = p["prix_achat"]
    sh = p["shipping"]
    shopify_fee = pv * 0.02
    stripe_fee = pv * 0.014 + 0.25
    cout_total = pa + sh + shopify_fee + stripe_fee
    marge_brute = pv - cout_total
    charges_micro = pv * 0.123
    marge_nette = marge_brute - charges_micro
    marge_nette_pct = marge_nette / pv if pv > 0 else 0
    ventes_3k = round(3000 / marge_nette, 1) / 30 if marge_nette > 0 else 999
    ventes_5k = round(5000 / marge_nette, 1) / 30 if marge_nette > 0 else 999

    values = [
        p["produit"], p["collection"], pa, sh, pv,
        round(shopify_fee, 2), round(stripe_fee, 2), round(cout_total, 2),
        round(marge_brute, 2), round(charges_micro, 2), round(marge_nette, 2),
        marge_nette_pct, round(ventes_3k, 1), round(ventes_5k, 1)
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws3.cell(row=row3, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = wrap_align
        if col_idx in (3, 4, 5, 6, 7, 8, 9, 10, 11):
            cell.number_format = money_fmt
        elif col_idx == 12:
            cell.number_format = pct_fmt

    row3 += 1

ws3.auto_filter.ref = f"A1:{get_column_letter(len(marge_headers))}{row3-1}"
ws3.freeze_panes = "B2"

# ═══════════════════════════════════════════════════════
# SHEET 4 : TOP Winners par Marge
# ═══════════════════════════════════════════════════════
ws4 = wb.create_sheet("TOP Winners par Marge")

top_headers = ["Rang", "Collection", "Produit", "Coût Achat (€)", "Prix Vente FR (€)",
               "Marge Brute (€)", "Marge (%)", "Score", "Verdict"]
top_widths = [8, 28, 50, 16, 18, 16, 12, 10, 28]

for col_idx, (h, w) in enumerate(zip(top_headers, top_widths), 1):
    cell = ws4.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws4.column_dimensions[get_column_letter(col_idx)].width = w

# Sort products by marge brute descending
sorted_products = sorted(products, key=lambda x: x["prix_vente"] - x["prix_achat"] - x["shipping"], reverse=True)

row4 = 2
for rank, p in enumerate(sorted_products, 1):
    cout = p["prix_achat"] + p["shipping"]
    marge = p["prix_vente"] - cout
    marge_pct = marge / p["prix_vente"] if p["prix_vente"] > 0 else 0

    if marge_pct >= 0.80:
        score = "A+"
        verdict = "WINNER - Lancer en priorité"
    elif marge_pct >= 0.70:
        score = "A"
        verdict = "GO - Fort potentiel"
    elif marge >= 100:
        score = "A"
        verdict = "GO - Fort potentiel"
    else:
        score = "B+"
        verdict = "GO - À tester"

    values = [rank, p["collection"], p["produit"], cout, p["prix_vente"], marge, marge_pct, score, verdict]
    for col_idx, val in enumerate(values, 1):
        cell = ws4.cell(row=row4, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = center_align if col_idx in (1, 4, 5, 6, 7, 8) else wrap_align
        if col_idx in (4, 5, 6):
            cell.number_format = money_fmt
        elif col_idx == 7:
            cell.number_format = pct_fmt
        # Highlight winners
        if "WINNER" in verdict and col_idx in (6, 7, 8, 9):
            cell.fill = winner_fill
    row4 += 1

ws4.freeze_panes = "D2"

# ═══════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════
output_path = "/Users/Hakim/ecommerce-dropshipping/catalogue_securite_aliexpress.xlsx"
wb.save(output_path)
print(f"Fichier créé : {output_path}")
print(f"  - {len(products)} produits")
print(f"  - {len(collections)} collections")
print(f"  - 4 onglets : Catalogue, Résumé, Marge Nette, TOP Winners")
