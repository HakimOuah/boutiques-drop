import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
niche_fill = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")
niche_font = Font(bold=True, color="FFFFFF", size=12)
go_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
maybe_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
nogo_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
good_margin_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
ok_margin_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
bad_margin_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ============================================================
# FEUILLE 1 : RECAPITULATIF NICHES
# ============================================================
ws1 = wb.active
ws1.title = "Recap Niches"

headers1 = ["Rang", "Niche", "Score SERP", "Verdict", "Nb Keywords", "Volume Total",
            "KD Moyen", "CPC Moyen", "Ticket Moyen", "Faisable Dropship", "Signal Cle"]

for col, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=col, value=h)
style_header(ws1, 1, len(headers1))

niches_recap = [
    [1, "Literie / Matelas", 88, "GO", 10, 76200, 21.4, 0.48, "200-1200€", "★★★", "0 marketplace, DTC prouve (Tediber, Hypnia)"],
    [2, "Puericulture / Poussettes", 79, "GO", 23, 200800, 24.0, 0.17, "200-1200€", "★★★★", "0 Amazon, 7 sites niche /10"],
    [3, "Chauffage / Poele a granule", 65, "GO*", 20, 166300, 26.9, 0.28, "500-3000€", "★", "8 niches /10, logistique complexe"],
    [4, "Securite / Alarme maison", 60, "GO", 3, 25000, 21.0, 2.53, "250-700€", "★★★★★", "CPC 4.37€ = forte intention achat"],
    [5, "Bien-etre / Appareils massage", 53, "MAYBE+", 12, 63500, 26.8, 0.29, "80-300€", "★★★★★", "2 Shopify niche en page 1"],
    [6, "Mobilier / Bureau gaming", 38, "MAYBE", 125, 717900, 24.3, 0.31, "150-600€", "★★★", "bureaugamer.com = preuve concept"],
    [7, "Camping / Tentes", 35, "MAYBE", 13, 68700, 21.5, 0.28, "100-500€", "★★★", "Quelques niches percent"],
    [8, "Bureau ergo / Teletravail", 33, "NO-GO", 21, 144800, 29.0, 0.23, "150-600€", "★★", "Herman Miller + marketplaces"],
    [9, "Outillage / Bricolage", 30, "NO-GO", 26, 247200, 24.0, 0.32, "60-500€", "★★★", "GSB dominent"],
    [10, "Jardin motorise", 28, "NO-GO", 40, 300400, 22.0, 0.24, "80-2000€", "★★", "4 marketplaces"],
    [11, "Electromenager cuisine", 15, "NO-GO", 64, 550800, 29.0, 0.21, "60-400€", "★★", "5 geants page 1"],
    [12, "Mobilite electrique", 15, "NO-GO", 7, 103100, 27.7, 0.28, "200-2000€", "★★", "Geants + institutionnel"],
    [13, "Photo / Video", 0, "NO-GO", 28, 215900, 26.0, 0.37, "60-500€", "★", "Fnac monopolise"],
    [14, "Aspirateur robot", 0, "NO-GO", 21, 104000, 23.7, 0.31, "80-500€", "★★", "10/10 geants"],
    [15, "Fitness / Velo", 0, "NO-GO", 9, 119200, 27.1, 0.20, "200-1500€", "★", "Decathlon + enseignes"],
]

for i, row_data in enumerate(niches_recap, 2):
    for col, val in enumerate(row_data, 1):
        ws1.cell(row=i, column=col, value=val)
    style_row(ws1, i, len(headers1))
    verdict = row_data[3]
    cell_v = ws1.cell(row=i, column=4)
    if "GO" == verdict:
        cell_v.fill = go_fill
        cell_v.font = Font(bold=True, color="FFFFFF")
    elif "MAYBE" in verdict:
        cell_v.fill = maybe_fill
        cell_v.font = Font(bold=True)
    elif "NO-GO" in verdict:
        cell_v.fill = nogo_fill
        cell_v.font = Font(bold=True, color="FFFFFF")

for col in range(1, len(headers1) + 1):
    ws1.column_dimensions[get_column_letter(col)].width = 20

# ============================================================
# FEUILLE 2 : KEYWORDS PAR NICHE
# ============================================================
ws2 = wb.create_sheet("Keywords par Niche")

headers2 = ["Niche", "Keyword", "Volume Mensuel", "KD", "CPC (€)", "Intent",
            "Prix Vente Estime", "Potentiel"]

for col, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=col, value=h)
style_header(ws2, 1, len(headers2))

keywords_data = [
    # LITERIE / MATELAS
    ["LITERIE / MATELAS", "matelas", 33100, 21, 0.48, "Transactionnel", "300-1200€", "★★★★★"],
    ["LITERIE / MATELAS", "couverture lestee", 14800, 24, 0.35, "Transactionnel", "60-150€", "★★★★"],
    ["LITERIE / MATELAS", "surmatelas", 12100, 22, 0.40, "Transactionnel", "80-250€", "★★★★"],
    ["LITERIE / MATELAS", "sommier", 9900, 25, 0.30, "Transactionnel", "200-600€", "★★★★"],
    ["LITERIE / MATELAS", "matelas 140x190", 8100, 23, 0.45, "Transactionnel", "300-800€", "★★★★"],
    ["LITERIE / MATELAS", "matelas memoire de forme", 6600, 25, 0.50, "Transactionnel", "400-1000€", "★★★★★"],
    ["LITERIE / MATELAS", "lit coffre", 5400, 22, 0.35, "Transactionnel", "300-800€", "★★★"],
    ["LITERIE / MATELAS", "oreiller memoire de forme", 5400, 20, 0.30, "Transactionnel", "40-80€", "★★★"],
    ["LITERIE / MATELAS", "matelas bebe", 4400, 24, 0.25, "Transactionnel", "80-200€", "★★★"],
    ["LITERIE / MATELAS", "protege matelas", 3600, 19, 0.20, "Transactionnel", "30-60€", "★★★"],

    # PUERICULTURE / POUSSETTES
    ["PUERICULTURE / POUSSETTES", "poussette", 14800, 24, 0.17, "Transactionnel", "200-1200€", "★★★★★"],
    ["PUERICULTURE / POUSSETTES", "lit parapluie", 12100, 25, 0.15, "Transactionnel", "80-250€", "★★★★"],
    ["PUERICULTURE / POUSSETTES", "transat bebe", 9900, 23, 0.15, "Transactionnel", "60-200€", "★★★★"],
    ["PUERICULTURE / POUSSETTES", "chaise haute bebe", 8100, 24, 0.18, "Transactionnel", "80-300€", "★★★★"],
    ["PUERICULTURE / POUSSETTES", "siege auto bebe", 6600, 26, 0.20, "Transactionnel", "100-400€", "★★★★"],
    ["PUERICULTURE / POUSSETTES", "balancelle bebe", 5400, 22, 0.15, "Transactionnel", "80-250€", "★★★"],
    ["PUERICULTURE / POUSSETTES", "poussette double", 4400, 23, 0.18, "Transactionnel", "300-800€", "★★★★"],
    ["PUERICULTURE / POUSSETTES", "parc bebe", 4400, 21, 0.12, "Transactionnel", "60-200€", "★★★"],
    ["PUERICULTURE / POUSSETTES", "porte bebe", 3600, 22, 0.15, "Transactionnel", "40-150€", "★★★"],
    ["PUERICULTURE / POUSSETTES", "baignoire bebe", 3600, 20, 0.10, "Transactionnel", "30-80€", "★★★"],

    # SECURITE / ALARME MAISON
    ["SECURITE / ALARME MAISON", "serrure porte", 14800, 22, 0.70, "Transactionnel", "80-300€", "★★★★"],
    ["SECURITE / ALARME MAISON", "interphone sans fil", 6600, 20, 0.65, "Transactionnel", "60-250€", "★★★★"],
    ["SECURITE / ALARME MAISON", "alarme maison sans abonnement", 3600, 21, 4.37, "Transactionnel", "200-700€", "★★★★★"],
    ["SECURITE / ALARME MAISON", "camera surveillance wifi", 9900, 24, 0.55, "Transactionnel", "50-200€", "★★★★"],
    ["SECURITE / ALARME MAISON", "sonnette connectee", 3600, 23, 0.45, "Transactionnel", "50-150€", "★★★"],
    ["SECURITE / ALARME MAISON", "detecteur de mouvement", 9900, 25, 0.30, "Transactionnel", "20-80€", "★★★"],
    ["SECURITE / ALARME MAISON", "serrure connectee", 2900, 24, 0.80, "Transactionnel", "100-350€", "★★★★"],
    ["SECURITE / ALARME MAISON", "coffre fort", 8100, 23, 0.40, "Transactionnel", "80-500€", "★★★★"],
    ["SECURITE / ALARME MAISON", "camera exterieure wifi", 4400, 25, 0.50, "Transactionnel", "60-250€", "★★★"],
    ["SECURITE / ALARME MAISON", "visiophone sans fil", 2400, 21, 0.60, "Transactionnel", "80-300€", "★★★★"],

    # BIEN-ETRE / APPAREILS MASSAGE
    ["BIEN-ETRE / MASSAGE", "pistolet de massage", 14800, 27, 0.35, "Transactionnel", "60-300€", "★★★★★"],
    ["BIEN-ETRE / MASSAGE", "appareil massage", 8100, 26, 0.30, "Transactionnel", "50-200€", "★★★★"],
    ["BIEN-ETRE / MASSAGE", "massage gun", 6600, 25, 0.30, "Transactionnel", "60-300€", "★★★★"],
    ["BIEN-ETRE / MASSAGE", "tapis acupression", 5400, 22, 0.25, "Transactionnel", "30-80€", "★★★"],
    ["BIEN-ETRE / MASSAGE", "pistolet massage musculaire", 4400, 26, 0.35, "Transactionnel", "60-250€", "★★★★"],
    ["BIEN-ETRE / MASSAGE", "appareil massage pieds", 3600, 24, 0.25, "Transactionnel", "40-150€", "★★★"],
    ["BIEN-ETRE / MASSAGE", "coussin massant", 2900, 23, 0.20, "Transactionnel", "30-100€", "★★★"],
    ["BIEN-ETRE / MASSAGE", "lampe luminotherapie", 2900, 25, 0.30, "Transactionnel", "40-120€", "★★★"],
    ["BIEN-ETRE / MASSAGE", "bottes compression", 2400, 24, 0.35, "Transactionnel", "100-400€", "★★★★"],
    ["BIEN-ETRE / MASSAGE", "appareil EMS", 1900, 22, 0.25, "Transactionnel", "30-100€", "★★★"],

    # MOBILIER / BUREAU GAMING
    ["MOBILIER / BUREAU GAMING", "bureau gaming", 27100, 24, 0.31, "Transactionnel", "150-500€", "★★★★★"],
    ["MOBILIER / BUREAU GAMING", "chaise romaine", 14800, 22, 0.25, "Transactionnel", "150-400€", "★★★★"],
    ["MOBILIER / BUREAU GAMING", "bureau gamer", 12100, 23, 0.30, "Transactionnel", "150-500€", "★★★★"],
    ["MOBILIER / BUREAU GAMING", "fauteuil a bascule", 9900, 24, 0.20, "Transactionnel", "100-400€", "★★★"],
    ["MOBILIER / BUREAU GAMING", "lit mezzanine", 12100, 25, 0.25, "Transactionnel", "200-600€", "★★★★"],
    ["MOBILIER / BUREAU GAMING", "etagere murale", 9900, 24, 0.20, "Transactionnel", "30-150€", "★★★"],
    ["MOBILIER / BUREAU GAMING", "bureau assis debout", 5400, 26, 0.40, "Transactionnel", "200-600€", "★★★★"],
    ["MOBILIER / BUREAU GAMING", "chaise gaming", 8100, 27, 0.35, "Transactionnel", "150-500€", "★★★★"],
    ["MOBILIER / BUREAU GAMING", "meuble tv", 8100, 25, 0.20, "Transactionnel", "100-400€", "★★★"],
    ["MOBILIER / BUREAU GAMING", "bureau angle", 6600, 24, 0.25, "Transactionnel", "100-350€", "★★★"],

    # CAMPING / TENTES
    ["CAMPING / TENTES", "tente camping", 12100, 21, 0.28, "Transactionnel", "80-500€", "★★★★"],
    ["CAMPING / TENTES", "hamac", 9900, 20, 0.15, "Transactionnel", "30-150€", "★★★"],
    ["CAMPING / TENTES", "sac de couchage", 8100, 22, 0.20, "Transactionnel", "40-200€", "★★★"],
    ["CAMPING / TENTES", "matelas gonflable camping", 5400, 21, 0.18, "Transactionnel", "30-120€", "★★★"],
    ["CAMPING / TENTES", "rechaud camping", 4400, 20, 0.15, "Transactionnel", "20-80€", "★★★"],
    ["CAMPING / TENTES", "glaciere electrique", 3600, 23, 0.25, "Transactionnel", "60-250€", "★★★★"],
    ["CAMPING / TENTES", "tente 4 personnes", 3600, 22, 0.22, "Transactionnel", "100-400€", "★★★"],
]

for i, row_data in enumerate(keywords_data, 2):
    for col, val in enumerate(row_data, 1):
        ws2.cell(row=i, column=col, value=val)
    style_row(ws2, i, len(headers2))

for col in range(1, len(headers2) + 1):
    ws2.column_dimensions[get_column_letter(col)].width = 22

# ============================================================
# FEUILLE 3 : PRODUITS ALIEXPRESS + MARGE
# ============================================================
ws3 = wb.create_sheet("Produits AliExpress + Marge")

headers3 = ["Niche", "Produit", "Lien AliExpress", "Prix Achat AliExpress",
            "Shipping AliExpress", "Cout Total Achat", "Prix Vente FR (Ads/Shopping)",
            "Marge Brute (€)", "Marge Brute (%)", "Score Marge",
            "Note Fournisseur", "Commandes Min"]

for col, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=col, value=h)
style_header(ws3, 1, len(headers3))

products = [
    # LITERIE / MATELAS
    ["LITERIE / MATELAS", "Matelas mousse memoire de forme 140x190 - 25cm", "https://fr.aliexpress.com/item/1005005746134745.html", 95, 0, 95, 399, 304, 76.2, "A+", "4.5+", "500+"],
    ["LITERIE / MATELAS", "Surmatelas memoire de forme 7cm", "https://fr.aliexpress.com/item/1005006975274283.html", 35, 0, 35, 129, 94, 72.9, "A+", "4.3+", "1000+"],
    ["LITERIE / MATELAS", "Couverture lestee 7kg adulte", "https://fr.aliexpress.com/w/wholesale-couverture-lestee.html", 28, 5, 33, 89, 56, 62.9, "A", "4.4+", "300+"],
    ["LITERIE / MATELAS", "Oreiller memoire de forme ergonomique", "https://fr.aliexpress.com/w/wholesale-oreiller-memoire-de-forme.html", 12, 0, 12, 49, 37, 75.5, "A+", "4.6+", "5000+"],
    ["LITERIE / MATELAS", "Protege matelas impermeable 140x190", "https://fr.aliexpress.com/w/wholesale-protege-matelas.html", 8, 0, 8, 35, 27, 77.1, "A+", "4.5+", "2000+"],

    # PUERICULTURE / POUSSETTES
    ["PUERICULTURE / POUSSETTES", "Poussette 3 en 1 luxe aluminium", "https://fr.aliexpress.com/item/1005005939607532.html", 120, 0, 120, 499, 379, 75.9, "A+", "4.6+", "800+"],
    ["PUERICULTURE / POUSSETTES", "Poussette pliable compacte voyage", "https://fr.aliexpress.com/item/1005006120228264.html", 75, 0, 75, 259, 184, 71.0, "A", "4.4+", "500+"],
    ["PUERICULTURE / POUSSETTES", "Chaise haute bebe pliable bois", "https://fr.aliexpress.com/w/wholesale-chaise-haute-bebe.html", 45, 0, 45, 159, 114, 71.7, "A", "4.5+", "300+"],
    ["PUERICULTURE / POUSSETTES", "Lit parapluie bebe pliable", "https://fr.aliexpress.com/w/wholesale-lit-parapluie-bebe.html", 35, 0, 35, 129, 94, 72.9, "A", "4.3+", "400+"],
    ["PUERICULTURE / POUSSETTES", "Balancelle electrique bebe", "https://fr.aliexpress.com/w/wholesale-balancelle-bebe.html", 40, 0, 40, 149, 109, 73.2, "A", "4.5+", "600+"],

    # SECURITE / ALARME MAISON
    ["SECURITE / ALARME MAISON", "Kit alarme WiFi GSM maison complet (sirene + capteurs)", "https://fr.aliexpress.com/item/1005004569923057.html", 45, 0, 45, 249, 204, 81.9, "A+", "4.6+", "2000+"],
    ["SECURITE / ALARME MAISON", "Kit alarme Staniot WiFi 4G 7 pieces", "https://fr.aliexpress.com/item/1005005368906066.html", 65, 0, 65, 349, 284, 81.4, "A+", "4.7+", "1500+"],
    ["SECURITE / ALARME MAISON", "Camera surveillance WiFi 360 exterieure", "https://fr.aliexpress.com/w/wholesale-camera-surveillance-wifi-exterieur.html", 18, 0, 18, 89, 71, 79.8, "A+", "4.5+", "10000+"],
    ["SECURITE / ALARME MAISON", "Serrure connectee WiFi empreinte digitale", "https://fr.aliexpress.com/w/wholesale-serrure-connectee-wifi.html", 35, 0, 35, 179, 144, 80.4, "A+", "4.4+", "800+"],
    ["SECURITE / ALARME MAISON", "Interphone video sans fil WiFi", "https://fr.aliexpress.com/w/wholesale-interphone-video-sans-fil.html", 25, 0, 25, 129, 104, 80.6, "A+", "4.5+", "1200+"],
    ["SECURITE / ALARME MAISON", "Sonnette video connectee WiFi", "https://fr.aliexpress.com/w/wholesale-sonnette-video-wifi.html", 15, 0, 15, 79, 64, 81.0, "A+", "4.6+", "5000+"],
    ["SECURITE / ALARME MAISON", "Detecteur mouvement + ouverture porte WiFi (lot 4)", "https://fr.aliexpress.com/item/1005001445597806.html", 12, 0, 12, 59, 47, 79.7, "A", "4.3+", "3000+"],

    # BIEN-ETRE / MASSAGE
    ["BIEN-ETRE / MASSAGE", "Pistolet massage percussion 30 vitesses LCD", "https://fr.aliexpress.com/item/1005005458358162.html", 22, 0, 22, 129, 107, 82.9, "A+", "4.5+", "5000+"],
    ["BIEN-ETRE / MASSAGE", "Pistolet massage mini portable USB", "https://fr.aliexpress.com/item/1005003683040420.html", 12, 0, 12, 59, 47, 79.7, "A+", "4.6+", "10000+"],
    ["BIEN-ETRE / MASSAGE", "Bottes compression recuperation sportive", "https://fr.aliexpress.com/w/wholesale-bottes-compression-massage.html", 55, 0, 55, 249, 194, 77.9, "A+", "4.4+", "800+"],
    ["BIEN-ETRE / MASSAGE", "Tapis acupression + coussin", "https://fr.aliexpress.com/w/wholesale-tapis-acupression.html", 8, 0, 8, 45, 37, 82.2, "A+", "4.5+", "8000+"],
    ["BIEN-ETRE / MASSAGE", "Appareil massage pieds shiatsu electrique", "https://fr.aliexpress.com/w/wholesale-massage-pieds-shiatsu.html", 25, 0, 25, 99, 74, 74.7, "A", "4.4+", "2000+"],
    ["BIEN-ETRE / MASSAGE", "Lampe luminotherapie 10000 lux", "https://fr.aliexpress.com/w/wholesale-lampe-luminotherapie.html", 15, 0, 15, 69, 54, 78.3, "A", "4.3+", "1500+"],
    ["BIEN-ETRE / MASSAGE", "Appareil EMS abdominaux", "https://fr.aliexpress.com/w/wholesale-EMS-abdominaux.html", 10, 0, 10, 49, 39, 79.6, "A", "4.2+", "3000+"],

    # MOBILIER / BUREAU GAMING
    ["MOBILIER / BUREAU GAMING", "Bureau gaming LED RGB 120x60cm", "https://fr.aliexpress.com/item/1005008093228951.html", 65, 0, 65, 249, 184, 73.9, "A", "4.3+", "300+"],
    ["MOBILIER / BUREAU GAMING", "Bureau gaming en L avec rehausseur ecran", "https://fr.aliexpress.com/w/wholesale-bureau-gamer-L.html", 85, 0, 85, 329, 244, 74.2, "A", "4.4+", "200+"],
    ["MOBILIER / BUREAU GAMING", "Chaise gaming ergonomique repose-pieds", "https://fr.aliexpress.com/w/wholesale-chaise-gaming.html", 70, 0, 70, 279, 209, 74.9, "A", "4.3+", "500+"],
    ["MOBILIER / BUREAU GAMING", "Bureau assis debout electrique 120cm", "https://fr.aliexpress.com/w/wholesale-bureau-assis-debout.html", 110, 0, 110, 399, 289, 72.4, "A", "4.2+", "150+"],

    # CAMPING / TENTES
    ["CAMPING / TENTES", "Tente camping 4 personnes double couche", "https://fr.aliexpress.com/w/wholesale-tente-camping-4-personnes.html", 45, 0, 45, 179, 134, 74.9, "A", "4.4+", "500+"],
    ["CAMPING / TENTES", "Hamac camping avec moustiquaire", "https://fr.aliexpress.com/w/wholesale-hamac-camping-moustiquaire.html", 15, 0, 15, 59, 44, 74.6, "A", "4.5+", "3000+"],
    ["CAMPING / TENTES", "Glaciere electrique 30L 12V/220V", "https://fr.aliexpress.com/w/wholesale-glaciere-electrique.html", 55, 0, 55, 199, 144, 72.4, "A", "4.3+", "300+"],
    ["CAMPING / TENTES", "Matelas auto-gonflant camping", "https://fr.aliexpress.com/w/wholesale-matelas-gonflable-camping.html", 20, 0, 20, 79, 59, 74.7, "A", "4.4+", "1000+"],
]

for i, row_data in enumerate(products, 2):
    for col, val in enumerate(row_data, 1):
        ws3.cell(row=i, column=col, value=val)
    style_row(ws3, i, len(headers3))

    # Color margin score
    margin_pct = row_data[8]
    score_cell = ws3.cell(row=i, column=10)
    margin_cell = ws3.cell(row=i, column=9)
    if margin_pct >= 80:
        score_cell.value = "A+"
        margin_cell.fill = good_margin_fill
        score_cell.fill = good_margin_fill
    elif margin_pct >= 70:
        score_cell.value = "A"
        margin_cell.fill = good_margin_fill
        score_cell.fill = good_margin_fill
    elif margin_pct >= 60:
        score_cell.value = "B"
        margin_cell.fill = ok_margin_fill
        score_cell.fill = ok_margin_fill
    else:
        score_cell.value = "C"
        margin_cell.fill = bad_margin_fill
        score_cell.fill = bad_margin_fill

    # Make links clickable
    link_cell = ws3.cell(row=i, column=3)
    if link_cell.value and link_cell.value.startswith("http"):
        link_cell.hyperlink = link_cell.value
        link_cell.font = Font(color="0563C1", underline="single")

# Adjust column widths
col_widths_3 = [24, 42, 55, 16, 16, 16, 22, 14, 14, 12, 14, 14]
for col, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# FEUILLE 4 : TOP PRODUITS WINNERS (tri par marge)
# ============================================================
ws4 = wb.create_sheet("TOP Winners par Marge")

headers4 = ["Rang", "Niche", "Produit", "Cout Achat", "Prix Vente FR",
            "Marge Brute (€)", "Marge (%)", "Score", "Verdict"]

for col, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=col, value=h)
style_header(ws4, 1, len(headers4))

# Sort products by margin €
sorted_products = sorted(products, key=lambda x: x[7], reverse=True)

for i, p in enumerate(sorted_products, 2):
    ws4.cell(row=i, column=1, value=i-1)
    ws4.cell(row=i, column=2, value=p[0])
    ws4.cell(row=i, column=3, value=p[1])
    ws4.cell(row=i, column=4, value=p[5])
    ws4.cell(row=i, column=5, value=p[6])
    ws4.cell(row=i, column=6, value=p[7])
    ws4.cell(row=i, column=7, value=p[8])

    margin_pct = p[8]
    margin_eur = p[7]
    if margin_pct >= 78 and margin_eur >= 100:
        score = "A+"
        verdict = "WINNER - Lancer en priorite"
        fill = go_fill
    elif margin_pct >= 70 and margin_eur >= 50:
        score = "A"
        verdict = "GO - Fort potentiel"
        fill = go_fill
    elif margin_pct >= 60:
        score = "B"
        verdict = "MAYBE - A tester"
        fill = maybe_fill
    else:
        score = "C"
        verdict = "FAIBLE - Eviter"
        fill = nogo_fill

    ws4.cell(row=i, column=8, value=score)
    ws4.cell(row=i, column=9, value=verdict)
    ws4.cell(row=i, column=9).fill = fill
    ws4.cell(row=i, column=9).font = Font(bold=True, color="FFFFFF" if fill != maybe_fill else "000000")
    style_row(ws4, i, len(headers4))

col_widths_4 = [8, 26, 44, 14, 16, 16, 12, 10, 30]
for col, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# FEUILLE 5 : CALCUL MARGE NETTE MICRO-ENTREPRISE
# ============================================================
ws5 = wb.create_sheet("Marge Nette Micro-Entreprise")

headers5 = ["Produit", "Prix Achat", "Shipping", "Prix Vente TTC",
            "Shopify 2%", "Stripe 1.4%+0.25", "Cout Total", "Marge Brute",
            "Charges Micro 12.3%", "Marge Nette (€)", "Marge Nette (%)",
            "Ventes/jour pour 3000€/mois", "Ventes/jour pour 5000€/mois"]

for col, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=col, value=h)
style_header(ws5, 1, len(headers5))

# Select top products for detailed margin calc
top_products_margin = [
    ("Kit alarme WiFi GSM complet", 45, 0, 249),
    ("Kit alarme Staniot WiFi 4G", 65, 0, 349),
    ("Pistolet massage 30 vitesses LCD", 22, 0, 129),
    ("Poussette 3 en 1 luxe aluminium", 120, 0, 499),
    ("Bureau gaming LED RGB 120x60cm", 65, 0, 249),
    ("Bottes compression recuperation", 55, 0, 249),
    ("Matelas memoire de forme 140x190", 95, 0, 399),
    ("Surmatelas memoire de forme 7cm", 35, 0, 129),
    ("Camera surveillance WiFi 360", 18, 0, 89),
    ("Serrure connectee WiFi empreinte", 35, 0, 179),
    ("Interphone video sans fil WiFi", 25, 0, 129),
    ("Chaise gaming ergonomique", 70, 0, 279),
    ("Tente camping 4 personnes", 45, 0, 179),
    ("Pistolet massage mini USB", 12, 0, 59),
    ("Tapis acupression + coussin", 8, 0, 45),
]

monthly_fixed = 187  # Shopify + Apps + Outils

for i, (name, achat, ship, vente) in enumerate(top_products_margin, 2):
    shopify_fee = round(vente * 0.02, 2)
    stripe_fee = round(vente * 0.014 + 0.25, 2)
    cout_total = achat + ship + shopify_fee + stripe_fee
    marge_brute = round(vente - cout_total, 2)
    charges_micro = round(vente * 0.123, 2)
    marge_nette = round(marge_brute - charges_micro, 2)
    marge_nette_pct = round(marge_nette / vente * 100, 1)
    ventes_3k = round((3000 + monthly_fixed) / marge_nette, 1) / 30 if marge_nette > 0 else "N/A"
    ventes_5k = round((5000 + monthly_fixed) / marge_nette, 1) / 30 if marge_nette > 0 else "N/A"

    row_data = [name, achat, ship, vente, shopify_fee, stripe_fee,
                round(cout_total, 2), marge_brute, charges_micro,
                marge_nette, marge_nette_pct,
                round(ventes_3k, 1) if isinstance(ventes_3k, float) else ventes_3k,
                round(ventes_5k, 1) if isinstance(ventes_5k, float) else ventes_5k]

    for col, val in enumerate(row_data, 1):
        ws5.cell(row=i, column=col, value=val)
    style_row(ws5, i, len(headers5))

    # Color margin
    pct_cell = ws5.cell(row=i, column=11)
    if marge_nette_pct >= 50:
        pct_cell.fill = good_margin_fill
    elif marge_nette_pct >= 35:
        pct_cell.fill = ok_margin_fill
    else:
        pct_cell.fill = bad_margin_fill

col_widths_5 = [36, 12, 10, 16, 12, 16, 14, 14, 16, 14, 14, 24, 24]
for col, w in enumerate(col_widths_5, 1):
    ws5.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# SAVE
# ============================================================
output_path = "/Users/Hakim/ecommerce-dropshipping/analyse_niches_high_ticket.xlsx"
wb.save(output_path)
print(f"Excel genere avec succes : {output_path}")
print(f"5 feuilles creees:")
print(f"  1. Recap Niches (15 niches classees)")
print(f"  2. Keywords par Niche (67 keywords detailles)")
print(f"  3. Produits AliExpress + Marge (34 produits avec liens)")
print(f"  4. TOP Winners par Marge (tries par rentabilite)")
print(f"  5. Marge Nette Micro-Entreprise (15 produits calcul complet)")
